# app.py
import os
import sys
import uuid
import json
import re
import io
from datetime import datetime, timedelta
from flask import Flask, render_template, request, redirect, url_for, flash, session, send_from_directory, jsonify, send_file
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import bleach
from dotenv import load_dotenv
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# تحديد بيئة التشغيل
IS_VERCEL = os.environ.get('VERCEL', False)

load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'dev-secret-key-change-in-production')
app.permanent_session_lifetime = timedelta(days=7)

# إعدادات رفع الملفات
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

if IS_VERCEL:
    UPLOAD_FOLDER = '/tmp/uploads'
else:
    UPLOAD_FOLDER = os.path.join(BASE_DIR, 'static', 'uploads')

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp', 'pdf'}
MAX_CONTENT_LENGTH = 16 * 1024 * 1024  # 16MB

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_CONTENT_LENGTH

# إنشاء مجلدات الرفع
if not IS_VERCEL:
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    os.makedirs(os.path.join(UPLOAD_FOLDER, 'chalets'), exist_ok=True)
    os.makedirs(os.path.join(UPLOAD_FOLDER, 'categories'), exist_ok=True)
    os.makedirs(os.path.join(UPLOAD_FOLDER, 'national_id'), exist_ok=True)
    os.makedirs(os.path.join(UPLOAD_FOLDER, 'chalet_cards'), exist_ok=True)
else:
    os.makedirs('/tmp/uploads', exist_ok=True)
    os.makedirs('/tmp/uploads/chalets', exist_ok=True)
    os.makedirs('/tmp/uploads/categories', exist_ok=True)
    os.makedirs('/tmp/uploads/national_id', exist_ok=True)
    os.makedirs('/tmp/uploads/chalet_cards', exist_ok=True)

# --- دوال مساعدة ---
def is_valid_uuid(val):
    """التحقق من صحة UUID"""
    try:
        uuid.UUID(str(val))
        return True
    except ValueError:
        return False

# --- دوال الأمان ---
def sanitize_html(text):
    if text:
        return bleach.clean(text, tags=[], strip=True)
    return text

def allowed_file(filename):
    if not filename:
        return False
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_file_size(file):
    file.seek(0, os.SEEK_END)
    size = file.tell()
    file.seek(0)
    return size

def compress_image(file_path, max_size=(1200, 800), quality=85):
    try:
        img = Image.open(file_path)
        if img.mode in ('RGBA', 'LA'):
            background = Image.new('RGB', img.size, (255, 255, 255))
            background.paste(img, mask=img.split()[-1] if img.mode == 'RGBA' else None)
            img = background
        img.thumbnail(max_size, Image.Resampling.LANCZOS)
        img.save(file_path, optimize=True, quality=quality)
        return True
    except Exception as e:
        print(f"خطأ في ضغط الصورة: {e}")
        return False

def save_uploaded_file(file, subfolder=''):
    if not file or not file.filename or not allowed_file(file.filename):
        return None
    
    try:
        file_size = get_file_size(file)
        if file_size > MAX_CONTENT_LENGTH:
            flash('حجم الملف كبير جداً. الحد الأقصى 16 ميجابايت.', 'danger')
            return None
        
        ext = file.filename.rsplit('.', 1)[1].lower()
        filename = f"{uuid.uuid4().hex}.{ext}"
        
        upload_path = app.config['UPLOAD_FOLDER']
        if subfolder:
            upload_path = os.path.join(upload_path, subfolder)
            os.makedirs(upload_path, exist_ok=True)
        
        file_path = os.path.join(upload_path, filename)
        file.save(file_path)
        
        if ext in ['jpg', 'jpeg', 'png', 'gif', 'webp'] and file_size > 1 * 1024 * 1024:
            compress_image(file_path)
        
        if subfolder:
            return f"uploads/{subfolder}/{filename}"
        return f"uploads/{filename}"
        
    except Exception as e:
        print(f"خطأ في حفظ الملف: {e}")
        return None

def delete_uploaded_file(filepath):
    if not filepath:
        return False
    try:
        if filepath.startswith('uploads/'):
            filepath = filepath.replace('uploads/', '')
        upload_path = app.config['UPLOAD_FOLDER']
        full_path = os.path.join(upload_path, filepath)
        if os.path.exists(full_path):
            os.remove(full_path)
            return True
        return False
    except Exception as e:
        print(f"خطأ في حذف الملف: {e}")
        return False

def get_image_url(image_filename):
    if not image_filename:
        return url_for('static', filename='img/default_avatar.jpg')
    if image_filename.startswith('uploads/'):
        return url_for('static', filename=image_filename)
    if '/' in image_filename:
        return url_for('static', filename='uploads/' + image_filename)
    if image_filename.startswith('C') or image_filename.startswith('default'):
        return url_for('static', filename='img/' + image_filename)
    return url_for('static', filename='img/default_avatar.jpg')

@app.context_processor
def utility_processor():
    def get_image_url_dynamic(image_filename):
        return get_image_url(image_filename)
    return dict(get_image_url_dynamic=get_image_url_dynamic)

# --- استيراد قاعدة البيانات ---
import database as db

# --- دوال المصادقة ---
def login_required(role=None):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                flash('الرجاء تسجيل الدخول أولاً', 'warning')
                return redirect(url_for('login'))
            if role:
                user = db.get_user_by_id(session['user_id'])
                if user and user['role'] != role and user['role'] != 'admin':
                    flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
                    return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# --- الصفحات العامة ---

@app.route('/')
def index():
    chalets = db.get_all_chalets()[:6]
    categories = db.get_categories()
    return render_template('index.html', chalets=chalets, categories=categories, hero_image='C0.jpg')

@app.route('/about')
def about():
    categories = db.get_categories()
    return render_template('about.html', categories=categories)

@app.route('/contact')
def contact():
    return render_template('contact.html')

@app.route('/categories')
def categories():
    categories = db.get_categories()
    return render_template('categories.html', categories=categories)

@app.route('/category/<string:category_id>')
def category_chalets(category_id):
    if not is_valid_uuid(category_id):
        flash('التصنيف غير موجود', 'danger')
        return redirect(url_for('categories'))
    
    category = db.get_category_by_id(category_id)
    if not category:
        flash('التصنيف غير موجود', 'danger')
        return redirect(url_for('categories'))
    
    all_chalets = db.get_all_chalets()
    chalets = [c for c in all_chalets if c.get('category_id') == category_id]
    return render_template('category_chalets.html', category=category, chalets=chalets)

@app.route('/all-chalets')
def all_chalets():
    min_price = request.args.get('min_price', type=int)
    max_price = request.args.get('max_price', type=int)
    bedrooms = request.args.get('bedrooms', type=int)
    max_guests = request.args.get('max_guests', type=int)
    category_id = request.args.get('category_id')
    
    all_chalets = db.get_all_chalets()
    filtered_chalets = []
    for chalet in all_chalets:
        price = chalet['price_per_night']
        if min_price and price < min_price:
            continue
        if max_price and price > max_price:
            continue
        if bedrooms and chalet.get('bedrooms', 0) < bedrooms:
            continue
        if max_guests and chalet.get('max_guests', 0) < max_guests:
            continue
        if category_id and chalet.get('category_id') != category_id:
            continue
        filtered_chalets.append(chalet)
    
    categories = db.get_categories()
    return render_template('all_chalets.html', chalets=filtered_chalets, categories=categories)

@app.route('/chalet/<string:chalet_id>')
def chalet_detail(chalet_id):
    if not is_valid_uuid(chalet_id):
        flash('الشاليه غير موجود', 'danger')
        return redirect(url_for('all_chalets'))
    
    chalet = db.get_chalet_by_id(chalet_id)
    if not chalet:
        flash('الشاليه غير موجود', 'danger')
        return redirect(url_for('all_chalets'))
    
    booked_dates = db.get_all_booked_dates(chalet_id)
    return render_template('chalet_detail.html', chalet=chalet, booked_dates=booked_dates)

@app.route('/get-booked-dates/<string:chalet_id>')
def get_booked_dates(chalet_id):
    dates = db.get_all_booked_dates(chalet_id)
    return jsonify({'dates': dates})

# --- صفحات المصادقة ---

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        try:
            username = sanitize_html(request.form.get('username', '').strip())
            password = request.form.get('password', '').strip()
            confirm_password = request.form.get('confirm_password', '').strip()
            role = request.form.get('role', 'customer')
            name = sanitize_html(request.form.get('name', '').strip())
            email = sanitize_html(request.form.get('email', '').strip())
            phone = sanitize_html(request.form.get('phone', '').strip())
            national_id = sanitize_html(request.form.get('national_id', '').strip())
            
            if not all([username, password, name, phone, national_id]):
                flash('يرجى ملء جميع الحقول المطلوبة', 'danger')
                return render_template('signup.html')
            
            if password != confirm_password:
                flash('كلمات المرور غير متطابقة', 'danger')
                return render_template('signup.html')
            
            if len(password) < 6:
                flash('كلمة المرور يجب أن تكون 6 أحرف على الأقل', 'danger')
                return render_template('signup.html')
            
            
            existing = db.get_user_by_username(username)
            if existing:
                flash('اسم المستخدم موجود بالفعل', 'danger')
                return render_template('signup.html')
            
            national_id_image = save_uploaded_file(request.files.get('national_id_image'), 'national_id')
            
            user_data = {
                'username': username,
                'password': password,
                'role': role,
                'name': name,
                'email': email,
                'phone': phone,
                'national_id': national_id,
                'national_id_image': national_id_image
            }
            
            if role == 'owner':
                chalet_number = sanitize_html(request.form.get('chalet_number', '').strip())
                business_name = sanitize_html(request.form.get('business_name', '').strip())
                chalet_card_image = save_uploaded_file(request.files.get('chalet_card_image'), 'chalet_cards')
                
                if not chalet_number:
                    flash('رقم الشاليه مطلوب للمالكين', 'danger')
                    return render_template('signup.html')
                
                user_data.update({
                    'chalet_number': chalet_number,
                    'chalet_card_image': chalet_card_image,
                    'business_name': business_name
                })
            else:
                date_of_birth = request.form.get('date_of_birth', '').strip()
                address = sanitize_html(request.form.get('address', '').strip())
                emergency_contact = sanitize_html(request.form.get('emergency_contact', '').strip())
                user_data.update({
                    'date_of_birth': date_of_birth,
                    'address': address,
                    'emergency_contact': emergency_contact
                })
            
            user_id = db.create_user(user_data)
            
            if user_id:
                flash('تم التسجيل بنجاح! في انتظار موافقة المدير.', 'success')
                return redirect(url_for('login'))
            else:
                flash('حدث خطأ في إنشاء الحساب. يرجى المحاولة مرة أخرى.', 'danger')
                return render_template('signup.html')
            
        except Exception as e:
            print(f"❌ خطأ في التسجيل: {str(e)}")
            flash(f'حدث خطأ: {str(e)}', 'danger')
            return render_template('signup.html')
    
    return render_template('signup.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    session.clear()
    
    if request.method == 'POST':
        username = sanitize_html(request.form.get('username', '').strip())
        password = request.form.get('password', '').strip()
        
        if not username or not password:
            flash('الرجاء إدخال اسم المستخدم وكلمة المرور', 'danger')
            return render_template('login.html')
        
        try:
            user = db.get_user_by_username(username)
            
            if not user:
                flash('اسم المستخدم أو كلمة المرور غير صحيحة', 'danger')
                return render_template('login.html')
            
            if check_password_hash(user['password'], password):
                if user['role'] != 'admin' and user['status'] != 'approved':
                    if user['status'] == 'pending':
                        flash('حسابك في انتظار الموافقة من المدير', 'warning')
                    else:
                        flash('تم رفض حسابك. يرجى التواصل مع الدعم.', 'danger')
                    return render_template('login.html')
                
                session.permanent = True
                session['user_id'] = user['id']
                session['username'] = user['username']
                session['role'] = user['role']
                session['name'] = user['name']
                
                flash(f'مرحباً بعودتك, {user["name"]}!', 'success')
                return redirect(url_for('dashboard'))
            else:
                flash('اسم المستخدم أو كلمة المرور غير صحيحة', 'danger')
                
        except Exception as e:
            print(f"❌ خطأ في تسجيل الدخول: {e}")
            flash('حدث خطأ في تسجيل الدخول. يرجى المحاولة مرة أخرى.', 'danger')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('تم تسجيل الخروج بنجاح', 'info')
    return redirect(url_for('index'))

# --- لوحة التحكم ---

@app.route('/dashboard')
@login_required()
def dashboard():
    user = db.get_user_by_id(session['user_id'])
    if not user:
        session.clear()
        flash('انتهت الجلسة. الرجاء تسجيل الدخول مرة أخرى.', 'warning')
        return redirect(url_for('login'))
    
    role = user['role']
    user_id = user['id']
    
    if role == 'admin':
        return redirect(url_for('admin_dashboard'))
    elif role == 'owner':
        chalets = db.get_chalets_by_owner(user_id)
        bookings = db.get_bookings_by_owner(user_id)
        pending_count = len(db.get_pending_bookings_by_owner(user_id))
        return render_template('owner_dashboard.html', 
                               chalets=chalets, 
                               bookings=bookings,
                               user=user,
                               pending_count=pending_count)
    else:
        bookings = db.get_bookings_by_customer(user_id)
        return render_template('customer_dashboard.html', bookings=bookings, user=user)

# --- صفحات المالك ---

@app.route('/owner/add-chalet', methods=['GET', 'POST'])
@login_required(role='owner')
def owner_add_chalet():
    if request.method == 'POST':
        try:
            name_ar = sanitize_html(request.form.get('name_ar', '').strip())
            name_en = sanitize_html(request.form.get('name_en', '').strip())
            description_ar = sanitize_html(request.form.get('description_ar', '').strip())
            description_en = sanitize_html(request.form.get('description_en', '').strip())
            price = request.form.get('price', '0')
            category_id = request.form.get('category_id')
            location = sanitize_html(request.form.get('location', 'العين السخنة، السويس، مصر').strip())
            bedrooms = request.form.get('bedrooms', '2')
            bathrooms = request.form.get('bathrooms', '2')
            max_guests = request.form.get('max_guests', '6')
            amenities = sanitize_html(request.form.get('amenities', 'مسبح, واي فاي, تكييف').strip())
            
            if not name_ar or not description_ar:
                flash('الاسم والوصف بالعربية مطلوبان', 'danger')
                return render_template('owner_add_chalet.html', categories=db.get_categories())
            
            image = save_uploaded_file(request.files.get('image'), 'chalets') or 'C1.jpg'
            
            chalet_data = {
                'name_ar': name_ar,
                'name_en': name_en or name_ar,
                'description_ar': description_ar,
                'description_en': description_en or description_ar,
                'price_per_night': int(price) if price.isdigit() else 1000,
                'owner_id': session['user_id'],
                'category_id': category_id,
                'image': image,
                'location': location,
                'bedrooms': int(bedrooms) if bedrooms.isdigit() else 2,
                'bathrooms': int(bathrooms) if bathrooms.isdigit() else 2,
                'max_guests': int(max_guests) if max_guests.isdigit() else 6,
                'amenities': [a.strip() for a in amenities.split(',') if a.strip()]
            }
            
            db.create_chalet(chalet_data)
            flash('تم إضافة الشاليه بنجاح! في انتظار موافقة المدير.', 'success')
            return redirect(url_for('dashboard'))
        except Exception as e:
            flash(f'حدث خطأ: {str(e)}', 'danger')
            return render_template('owner_add_chalet.html', categories=db.get_categories())
    
    return render_template('owner_add_chalet.html', categories=db.get_categories())

@app.route('/owner/chalet/<string:chalet_id>/images', methods=['GET', 'POST'])
@login_required(role='owner')
def owner_chalet_images(chalet_id):
    if not is_valid_uuid(chalet_id):
        flash('الشاليه غير موجود', 'danger')
        return redirect(url_for('owner_dashboard'))
    
    chalet = db.get_chalet_by_id(chalet_id)
    if not chalet:
        flash('الشاليه غير موجود', 'danger')
        return redirect(url_for('owner_dashboard'))
    
    if chalet['owner_id'] != session['user_id']:
        flash('ليس لديك صلاحية لتعديل هذا الشاليه', 'danger')
        return redirect(url_for('owner_dashboard'))
    
    if request.method == 'POST':
        if 'image' in request.files:
            file = request.files['image']
            if file and file.filename and allowed_file(file.filename):
                image_path = save_uploaded_file(file, 'chalets')
                if image_path:
                    is_main = request.form.get('is_main') == 'on'
                    db.add_chalet_image(chalet_id, image_path, is_main)
                    flash('تم إضافة الصورة بنجاح', 'success')
                else:
                    flash('حدث خطأ في رفع الصورة', 'danger')
            else:
                flash('الرجاء اختيار صورة صالحة', 'danger')
        
        return redirect(url_for('owner_chalet_images', chalet_id=chalet_id))
    
    images = db.get_chalet_images(chalet_id)
    return render_template('owner_chalet_images.html', chalet=chalet, images=images)

@app.route('/owner/image/delete/<string:image_id>')
@login_required(role='owner')
def owner_delete_image(image_id):
    if not is_valid_uuid(image_id):
        flash('الصورة غير موجودة', 'danger')
        return redirect(url_for('owner_dashboard'))
    
    try:
        image = db.get_chalet_image_by_id(image_id)
        if not image:
            flash('الصورة غير موجودة', 'danger')
            return redirect(url_for('owner_dashboard'))
        
        # التحقق من الملكية
        chalet = db.get_chalet_by_id(image['chalet_id'])
        if not chalet or chalet['owner_id'] != session['user_id']:
            flash('ليس لديك صلاحية لحذف هذه الصورة', 'danger')
            return redirect(url_for('owner_dashboard'))
        
        delete_uploaded_file(image['image'])
        db.delete_chalet_image(image_id)
        flash('تم حذف الصورة بنجاح', 'success')
        return redirect(url_for('owner_chalet_images', chalet_id=image['chalet_id']))
    except Exception as e:
        flash(f'حدث خطأ: {str(e)}', 'danger')
        return redirect(url_for('owner_dashboard'))

@app.route('/owner/image/main/<string:image_id>')
@login_required(role='owner')
def owner_set_main_image(image_id):
    if not is_valid_uuid(image_id):
        flash('الصورة غير موجودة', 'danger')
        return redirect(url_for('owner_dashboard'))
    
    try:
        image = db.get_chalet_image_by_id(image_id)
        if not image:
            flash('الصورة غير موجودة', 'danger')
            return redirect(url_for('owner_dashboard'))
        
        chalet = db.get_chalet_by_id(image['chalet_id'])
        if not chalet or chalet['owner_id'] != session['user_id']:
            flash('ليس لديك صلاحية لتعديل هذه الصورة', 'danger')
            return redirect(url_for('owner_dashboard'))
        
        db.set_main_chalet_image(image_id, image['chalet_id'])
        flash('تم تعيين الصورة كصورة رئيسية', 'success')
        return redirect(url_for('owner_chalet_images', chalet_id=image['chalet_id']))
    except Exception as e:
        flash(f'حدث خطأ: {str(e)}', 'danger')
        return redirect(url_for('owner_dashboard'))

@app.route('/owner/chalet/<string:chalet_id>/dates', methods=['GET', 'POST'])
@login_required(role='owner')
def owner_chalet_dates(chalet_id):
    if not is_valid_uuid(chalet_id):
        flash('الشاليه غير موجود', 'danger')
        return redirect(url_for('owner_dashboard'))
    
    chalet = db.get_chalet_by_id(chalet_id)
    if not chalet or chalet['owner_id'] != session['user_id']:
        flash('ليس لديك صلاحية لتعديل هذا الشاليه', 'danger')
        return redirect(url_for('owner_dashboard'))
    
    if request.method == 'POST':
        action = request.form.get('action')
        date = request.form.get('date')
        reason = request.form.get('reason', '')
        
        if not date:
            flash('الرجاء اختيار تاريخ', 'danger')
            return redirect(url_for('owner_chalet_dates', chalet_id=chalet_id))
        
        if action == 'add':
            if db.add_owner_booked_date(chalet_id, date, reason):
                flash('تم إضافة التاريخ المحجوز بنجاح', 'success')
            else:
                flash('هذا التاريخ محجوز بالفعل', 'warning')
        elif action == 'remove':
            db.delete_owner_booked_date(chalet_id, date)
            flash('تم إزالة التاريخ المحجوز', 'success')
        
        return redirect(url_for('owner_chalet_dates', chalet_id=chalet_id))
    
    booked_dates = db.get_owner_booked_dates(chalet_id)
    booking_dates = db.get_booking_dates_for_chalet(chalet_id)
    
    return render_template('owner_chalet_dates.html', 
                         chalet=chalet, 
                         booked_dates=booked_dates,
                         booking_dates=booking_dates)

@app.route('/owner/bookings')
@login_required(role='owner')
def owner_bookings():
    user_id = session['user_id']
    filter_type = request.args.get('filter', 'all')
    
    pending_bookings = db.get_pending_bookings_by_owner(user_id)
    all_bookings = db.get_bookings_by_owner_all(user_id)
    
    if filter_type == 'pending':
        all_bookings = [b for b in all_bookings if b['status'] == 'pending']
    elif filter_type == 'confirmed':
        all_bookings = [b for b in all_bookings if b['status'] == 'confirmed']
    elif filter_type == 'completed':
        all_bookings = [b for b in all_bookings if b['status'] == 'completed']
    elif filter_type == 'cancelled':
        all_bookings = [b for b in all_bookings if b['status'] == 'cancelled']
    elif filter_type == 'rejected':
        all_bookings = [b for b in all_bookings if b['status'] == 'rejected']
    
    return render_template('owner_bookings.html', 
                         pending_bookings=pending_bookings,
                         all_bookings=all_bookings,
                         filter_type=filter_type)

@app.route('/owner/booking/approve/<string:booking_id>')
@login_required(role='owner')
def owner_approve_booking(booking_id):
    if db.update_booking_status_by_owner(booking_id, 'confirmed', session['user_id']):
        flash('تمت الموافقة على الحجز بنجاح', 'success')
    else:
        flash('حدث خطأ في الموافقة على الحجز', 'danger')
    return redirect(url_for('owner_bookings'))

@app.route('/owner/booking/reject/<string:booking_id>')
@login_required(role='owner')
def owner_reject_booking(booking_id):
    if db.update_booking_status_by_owner(booking_id, 'rejected', session['user_id']):
        flash('تم رفض الحجز', 'warning')
    else:
        flash('حدث خطأ في رفض الحجز', 'danger')
    return redirect(url_for('owner_bookings'))

@app.route('/owner/booking/cancel/<string:booking_id>')
@login_required(role='owner')
def owner_cancel_booking(booking_id):
    if db.update_booking_status_by_owner(booking_id, 'cancelled', session['user_id']):
        flash('تم إلغاء الحجز وإزالة التواريخ المحجوزة', 'success')
    else:
        flash('حدث خطأ في إلغاء الحجز', 'danger')
    return redirect(url_for('owner_bookings'))

@app.route('/owner/bookings/export')
@login_required(role='owner')
def owner_export_bookings():
    user_id = session['user_id']
    bookings = db.get_bookings_by_owner_all(user_id)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "حجوزاتي"
    
    headers = ['رقم الحجز', 'الشاليه', 'اسم العميل', 'رقم الهاتف', 'الرقم القومي',
               'تاريخ البداية', 'تاريخ النهاية', 'عدد الأيام', 'المبلغ (ج.م)', 'الحالة', 'تاريخ الحجز']
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a3c5e", end_color="1a3c5e", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    for row_idx, booking in enumerate(bookings, 2):
        status_map = {'pending': 'قيد الانتظار', 'confirmed': 'مؤكد', 'cancelled': 'ملغي', 'completed': 'مكتمل', 'rejected': 'مرفوض'}
        ws.cell(row=row_idx, column=1, value=booking['id'])
        ws.cell(row=row_idx, column=2, value=booking['chalet_name_ar'])
        ws.cell(row=row_idx, column=3, value=booking['customer_name'])
        ws.cell(row=row_idx, column=4, value=booking['customer_phone'])
        ws.cell(row=row_idx, column=5, value=booking.get('customer_national_id', ''))
        ws.cell(row=row_idx, column=6, value=booking['start_date'])
        ws.cell(row=row_idx, column=7, value=booking['end_date'])
        ws.cell(row=row_idx, column=8, value=booking['days'])
        ws.cell(row=row_idx, column=9, value=booking['amount'])
        ws.cell(row=row_idx, column=10, value=status_map.get(booking['status'], booking['status']))
        ws.cell(row=row_idx, column=11, value=booking['created_at'])
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'حجوزاتي_{datetime.now().strftime("%Y-%m-%d")}.xlsx')

# --- مسارات المدير ---

@app.route('/admin/dashboard')
@login_required(role='admin')
def admin_dashboard():
    stats = db.get_statistics()
    bookings = db.get_bookings()
    pending_users = db.get_pending_users()
    pending_chalets = db.get_pending_chalets()
    categories = db.get_categories()
    return render_template('admin_dashboard.html', 
                         stats=stats, bookings=bookings,
                         pending_users=pending_users, pending_chalets=pending_chalets,
                         categories=categories)

@app.route('/admin/chalets')
@login_required(role='admin')
def admin_chalets():
    pending_chalets = db.get_pending_chalets()
    return render_template('admin_chalets.html', pending_chalets=pending_chalets)

@app.route('/admin/chalet/approve/<string:chalet_id>')
@login_required(role='admin')
def admin_approve_chalet(chalet_id):
    db.approve_chalet(chalet_id, session['user_id'])
    flash('تمت الموافقة على الشاليه بنجاح', 'success')
    return redirect(url_for('admin_chalets'))

@app.route('/admin/chalet/reject/<string:chalet_id>')
@login_required(role='admin')
def admin_reject_chalet(chalet_id):
    db.reject_chalet(chalet_id)
    flash('تم رفض الشاليه', 'warning')
    return redirect(url_for('admin_chalets'))

@app.route('/admin/chalet/delete/<string:chalet_id>')
@login_required(role='admin')
def admin_delete_chalet(chalet_id):
    db.delete_chalet(chalet_id)
    flash('تم حذف الشاليه', 'success')
    return redirect(url_for('admin_chalets'))

@app.route('/admin/users')
@login_required(role='admin')
def admin_users():
    filter_type = request.args.get('filter', 'all')
    users = db.get_all_users()
    stats = db.get_user_statistics()
    
    if filter_type == 'pending':
        users = [u for u in users if u['status'] == 'pending']
    elif filter_type == 'approved':
        users = [u for u in users if u['status'] == 'approved']
    elif filter_type == 'rejected':
        users = [u for u in users if u['status'] == 'rejected']
    
    return render_template('admin_users.html', users=users, stats=stats, filter_type=filter_type)

@app.route('/admin/user/<string:user_id>')
@login_required(role='admin')
def admin_user_detail(user_id):
    if not is_valid_uuid(user_id):
        flash('المستخدم غير موجود', 'danger')
        return redirect(url_for('admin_users'))
    
    user = db.get_user_by_id(user_id)
    if not user:
        flash('المستخدم غير موجود', 'danger')
        return redirect(url_for('admin_users'))
    
    chalets_count = len(db.get_chalets_by_owner(user_id)) if user['role'] == 'owner' else 0
    return render_template('admin_user_detail.html', user=user, chalets_count=chalets_count)

@app.route('/admin/user/edit/<string:user_id>', methods=['GET', 'POST'])
@login_required(role='admin')
def admin_user_edit(user_id):
    if not is_valid_uuid(user_id):
        flash('المستخدم غير موجود', 'danger')
        return redirect(url_for('admin_users'))
    
    user = db.get_user_by_id(user_id)
    if not user:
        flash('المستخدم غير موجود', 'danger')
        return redirect(url_for('admin_users'))
    
    if user['role'] == 'admin' and user['id'] != session['user_id']:
        flash('لا يمكن تعديل بيانات مدير آخر', 'danger')
        return redirect(url_for('admin_users'))
    
    if request.method == 'POST':
        try:
            name = sanitize_html(request.form.get('name', '').strip())
            email = sanitize_html(request.form.get('email', '').strip())
            phone = sanitize_html(request.form.get('phone', '').strip())
            national_id = sanitize_html(request.form.get('national_id', '').strip())
            
            if not name:
                flash('الاسم مطلوب', 'danger')
                return render_template('admin_user_edit.html', user=user)
            
            data = {
                'name': name,
                'email': email,
                'phone': phone,
                'national_id': national_id
            }
            
            db.update_user(user_id, data)
            flash('تم تحديث بيانات المستخدم بنجاح', 'success')
            return redirect(url_for('admin_users'))
        except Exception as e:
            flash(f'حدث خطأ: {str(e)}', 'danger')
            return render_template('admin_user_edit.html', user=user)
    
    return render_template('admin_user_edit.html', user=user)

@app.route('/admin/user/upgrade/<string:user_id>', methods=['GET', 'POST'])
@login_required(role='admin')
def admin_user_upgrade(user_id):
    if not is_valid_uuid(user_id):
        flash('المستخدم غير موجود', 'danger')
        return redirect(url_for('admin_users'))
    
    user = db.get_user_by_id(user_id)
    if not user:
        flash('المستخدم غير موجود', 'danger')
        return redirect(url_for('admin_users'))
    
    if user['role'] == 'admin' and user['id'] != session['user_id']:
        flash('لا يمكن تغيير دور مدير آخر', 'danger')
        return redirect(url_for('admin_users'))
    
    if request.method == 'POST':
        new_role = request.form.get('new_role')
        
        if new_role not in ['customer', 'owner', 'admin']:
            flash('دور غير صحيح', 'danger')
            return redirect(url_for('admin_users'))
        
        if user['role'] == 'admin' and new_role != 'admin':
            stats = db.get_user_statistics()
            if stats['admins'] <= 1:
                flash('لا يمكن تنزيل المدير الأخير', 'danger')
                return redirect(url_for('admin_users'))
        
        if db.update_user_role(user_id, new_role):
            flash(f'تم تحديث دور المستخدم إلى {new_role} بنجاح', 'success')
        else:
            flash('حدث خطأ في تحديث الدور', 'danger')
        
        return redirect(url_for('admin_users'))
    
    return render_template('admin_user_upgrade.html', user=user)

@app.route('/admin/user/delete/<string:user_id>')
@login_required(role='admin')
def admin_user_delete(user_id):
    if not is_valid_uuid(user_id):
        flash('المستخدم غير موجود', 'danger')
        return redirect(url_for('admin_users'))
    
    if user_id == session['user_id']:
        flash('لا يمكن حذف حسابك الخاص', 'danger')
        return redirect(url_for('admin_users'))
    
    user = db.get_user_by_id(user_id)
    if not user:
        flash('المستخدم غير موجود', 'danger')
        return redirect(url_for('admin_users'))
    
    if user['role'] == 'admin':
        flash('لا يمكن حذف مدير', 'danger')
        return redirect(url_for('admin_users'))
    
    if db.delete_user(user_id):
        flash(f'تم حذف المستخدم {user["name"]} بنجاح', 'success')
    else:
        flash('حدث خطأ في حذف المستخدم', 'danger')
    
    return redirect(url_for('admin_users'))

@app.route('/admin/user/approve/<string:user_id>')
@login_required(role='admin')
def admin_approve_user(user_id):
    user = db.get_user_by_id(user_id)
    if not user:
        flash('المستخدم غير موجود', 'danger')
        return redirect(url_for('admin_users'))
    
    db.approve_user(user_id, session['user_id'])
    flash(f'تمت الموافقة على المستخدم {user["name"]}', 'success')
    return redirect(url_for('admin_users'))

@app.route('/admin/user/reject/<string:user_id>')
@login_required(role='admin')
def admin_reject_user(user_id):
    user = db.get_user_by_id(user_id)
    if not user:
        flash('المستخدم غير موجود', 'danger')
        return redirect(url_for('admin_users'))
    
    db.reject_user(user_id)
    flash(f'تم رفض المستخدم {user["name"]}', 'warning')
    return redirect(url_for('admin_users'))

@app.route('/admin/bookings')
@login_required(role='admin')
def admin_bookings():
    filter_type = request.args.get('filter', 'all')
    bookings = db.get_bookings()
    
    if filter_type == 'pending':
        bookings = [b for b in bookings if b['status'] == 'pending']
    elif filter_type == 'confirmed':
        bookings = [b for b in bookings if b['status'] == 'confirmed']
    elif filter_type == 'completed':
        bookings = [b for b in bookings if b['status'] == 'completed']
    elif filter_type == 'cancelled':
        bookings = [b for b in bookings if b['status'] == 'cancelled']
    elif filter_type == 'rejected':
        bookings = [b for b in bookings if b['status'] == 'rejected']
    
    return render_template('admin_bookings.html', bookings=bookings, filter_type=filter_type)

@app.route('/admin/bookings/export')
@login_required(role='admin')
def admin_export_bookings():
    bookings = db.get_bookings()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "الحجوزات"
    
    headers = ['رقم الحجز', 'الشاليه', 'اسم العميل', 'رقم الهاتف', 'الرقم القومي',
               'تاريخ البداية', 'تاريخ النهاية', 'عدد الأيام', 'المبلغ (ج.م)',
               'الحالة', 'اسم المالك', 'هاتف المالك', 'تاريخ الحجز']
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a3c5e", end_color="1a3c5e", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    for row_idx, booking in enumerate(bookings, 2):
        status_map = {'pending': 'قيد الانتظار', 'confirmed': 'مؤكد', 'cancelled': 'ملغي', 'completed': 'مكتمل', 'rejected': 'مرفوض'}
        ws.cell(row=row_idx, column=1, value=booking['id'])
        ws.cell(row=row_idx, column=2, value=booking['chalet_name_ar'])
        ws.cell(row=row_idx, column=3, value=booking['customer_name'])
        ws.cell(row=row_idx, column=4, value=booking['customer_phone'])
        ws.cell(row=row_idx, column=5, value=booking.get('customer_national_id', ''))
        ws.cell(row=row_idx, column=6, value=booking['start_date'])
        ws.cell(row=row_idx, column=7, value=booking['end_date'])
        ws.cell(row=row_idx, column=8, value=booking['days'])
        ws.cell(row=row_idx, column=9, value=booking['amount'])
        ws.cell(row=row_idx, column=10, value=status_map.get(booking['status'], booking['status']))
        ws.cell(row=row_idx, column=11, value=booking.get('owner_name', ''))
        ws.cell(row=row_idx, column=12, value=booking.get('owner_phone', ''))
        ws.cell(row=row_idx, column=13, value=booking['created_at'])
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    total_row = len(bookings) + 2
    ws.cell(row=total_row, column=8, value="الإجمالي:")
    ws.cell(row=total_row, column=9, value=f"=SUM(I2:I{len(bookings)+1})")
    ws.cell(row=total_row, column=8).font = Font(bold=True)
    ws.cell(row=total_row, column=9).font = Font(bold=True)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'الحجوزات_{datetime.now().strftime("%Y-%m-%d")}.xlsx')

@app.route('/admin/booking/approve/<string:booking_id>')
@login_required(role='admin')
def admin_approve_booking(booking_id):
    db.update_booking_status(booking_id, 'confirmed')
    flash('تمت الموافقة على الحجز', 'success')
    return redirect(url_for('admin_bookings'))

@app.route('/admin/booking/cancel/<string:booking_id>')
@login_required(role='admin')
def admin_cancel_booking(booking_id):
    db.update_booking_status(booking_id, 'cancelled')
    flash('تم إلغاء الحجز', 'warning')
    return redirect(url_for('admin_bookings'))

@app.route('/admin/booking/complete/<string:booking_id>')
@login_required(role='admin')
def admin_complete_booking(booking_id):
    db.update_booking_status(booking_id, 'completed')
    flash('تم إكمال الحجز', 'success')
    return redirect(url_for('admin_bookings'))

# --- صفحات العميل ---

@app.route('/book/<string:chalet_id>', methods=['GET', 'POST'])
@login_required(role='customer')
def book_chalet(chalet_id):
    if not is_valid_uuid(chalet_id):
        flash('الشاليه غير موجود', 'danger')
        return redirect(url_for('all_chalets'))
    
    chalet = db.get_chalet_by_id(chalet_id)
    if not chalet:
        flash('الشاليه غير موجود', 'danger')
        return redirect(url_for('all_chalets'))
    
    if request.method == 'POST':
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date')
        
        if not start_date or not end_date:
            flash('الرجاء اختيار تاريخ البداية والنهاية', 'danger')
            return render_template('book.html', chalet=chalet)
        
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d')
            end = datetime.strptime(end_date, '%Y-%m-%d')
            days = (end - start).days
            
            if days <= 0:
                flash('تاريخ النهاية يجب أن يكون بعد تاريخ البداية', 'danger')
                return render_template('book.html', chalet=chalet)
            
            if not db.check_availability_with_owner_dates(chalet_id, start_date, end_date):
                flash('هذه التواريخ غير متاحة. يرجى اختيار تواريخ أخرى.', 'danger')
                return render_template('book.html', chalet=chalet)
            
            total_amount = days * chalet['price_per_night']
            
            booking_data = {
                'chalet_id': chalet_id,
                'customer_id': session['user_id'],
                'start_date': start_date,
                'end_date': end_date,
                'amount': total_amount,
                'days': days,
                'payment_method': 'bank_transfer'
            }
            
            booking_id = db.create_booking(booking_data)
            
            flash(f'تم تقديم طلب الحجز بنجاح! في انتظار موافقة المالك. المبلغ: {total_amount} جنيه مصري لـ {days} ليلة.', 'success')
            return redirect(url_for('dashboard'))
            
        except Exception as e:
            flash(f'حدث خطأ: {str(e)}', 'danger')
            return render_template('book.html', chalet=chalet)
    
    booked_dates = db.get_all_booked_dates(chalet_id)
    return render_template('book.html', chalet=chalet, booked_dates=booked_dates)

@app.route('/booking/<string:booking_id>')
@login_required()
def booking_detail(booking_id):
    if not is_valid_uuid(booking_id):
        flash('الحجز غير موجود', 'danger')
        return redirect(url_for('dashboard'))
    
    user = db.get_user_by_id(session['user_id'])
    booking = db.get_booking_by_id(booking_id)
    
    if not booking:
        flash('الحجز غير موجود', 'danger')
        return redirect(url_for('dashboard'))
    
    is_admin = user['role'] == 'admin'
    is_owner = False
    is_customer = False
    
    if user['role'] == 'owner':
        chalet = db.get_chalet_by_id(booking['chalet_id'])
        if chalet and chalet['owner_id'] == user['id']:
            is_owner = True
    
    if user['role'] == 'customer' and booking['customer_id'] == user['id']:
        is_customer = True
    
    if not (is_admin or is_owner or is_customer):
        flash('ليس لديك صلاحية لعرض هذا الحجز', 'danger')
        return redirect(url_for('dashboard'))
    
    chalet = db.get_chalet_by_id(booking['chalet_id'])
    customer = db.get_user_by_id(booking['customer_id'])
    owner = db.get_user_by_id(chalet['owner_id']) if chalet else None
    
    return render_template('booking_detail.html', 
                         booking=booking, chalet=chalet,
                         customer=customer, owner=owner,
                         is_admin=is_admin, is_owner=is_owner, is_customer=is_customer)

# --- خدمة الملفات ---

@app.route('/static/uploads/<path:filename>')
def uploaded_file(filename):
    try:
        return send_from_directory(app.config['UPLOAD_FOLDER'], filename)
    except Exception as e:
        return send_from_directory(app.config['UPLOAD_FOLDER'], 'default_image.jpg')

# --- معالجة الأخطاء ---

@app.errorhandler(404)
def not_found_error(error):
    flash('الصفحة غير موجودة', 'danger')
    return redirect(url_for('index'))

@app.errorhandler(500)
def internal_error(error):
    flash('حدث خطأ داخلي. الرجاء المحاولة مرة أخرى.', 'danger')
    return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=4444)
